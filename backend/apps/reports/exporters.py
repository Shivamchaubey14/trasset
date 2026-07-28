"""
CSV and XLSX export (FR-10.2).

Memory is the constraint here: the register is specified to hold 100k+ assets
(NFR-5), so neither format may build the whole file in memory.

* **CSV** streams row by row through ``StreamingHttpResponse``. Nothing larger
  than a single row is ever held.
* **XLSX** uses openpyxl's ``write_only`` workbook, which flushes rows to a
  temporary file as they are appended. The zip container has to be finalised
  before the first byte can be sent, so the response is served from that temp
  file rather than streamed from the generator — memory stays flat either way.

PDF export is deliberately not implemented; it is deferred to v1.1.
"""
import csv
import datetime
import tempfile
from decimal import Decimal

from django.http import FileResponse, StreamingHttpResponse
from django.utils import timezone

CSV_CONTENT_TYPE = "text/csv; charset=utf-8"
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def filename_for(report, extension: str) -> str:
    stamp = timezone.now().strftime("%Y%m%d-%H%M")
    return f"trasset-{report.key}-{stamp}.{extension}"


class Echo:
    """A file-like object whose ``write`` returns the value it was given.

    ``csv.writer`` writes to it, and the generator yields whatever comes back —
    the standard trick for streaming CSV without an intermediate buffer.
    """

    def write(self, value):
        return value


def _csv_value(value):
    """Render a value for CSV — dates ISO, decimals plain, None as blank."""
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    return value


def stream_csv(report) -> StreamingHttpResponse:
    """Stream a report as CSV, one row at a time."""
    writer = csv.writer(Echo())

    def generate():
        # A BOM so Excel opens UTF-8 correctly on Windows — without it,
        # accented names and the rupee sign come out as mojibake.
        yield "﻿"
        yield writer.writerow(report.headers())

        for row in report.rows():
            yield writer.writerow(
                [_csv_value(row[column.key]) for column in report.columns]
            )

    response = StreamingHttpResponse(generate(), content_type=CSV_CONTENT_TYPE)
    response["Content-Disposition"] = (
        f'attachment; filename="{filename_for(report, "csv")}"'
    )
    return response


def _xlsx_value(value):
    """openpyxl handles dates and numbers natively; everything else is text."""
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _as_number(value):
    """Turn a decimal-ish string into a float, leaving anything else alone."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        try:
            return float(Decimal(value))
        except Exception:  # noqa: BLE001 - not a number, keep it as text
            return value
    return value


def build_xlsx(report) -> FileResponse:
    """Write a report to a temporary XLSX and serve it."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=report.title[:31] or "Report")

    # Header row, styled in Ink to match the brand.
    from openpyxl.cell import WriteOnlyCell

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", start_color="253D4E")
    header_cells = []
    for column in report.columns:
        cell = WriteOnlyCell(sheet, value=column.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    # Column widths and number formats have to be set before rows are written
    # in write-only mode.
    for index, column in enumerate(report.columns, start=1):
        letter = get_column_letter(index)
        if column.kind == "money":
            width = 16
        elif column.kind == "date":
            width = 14
        elif column.kind == "number":
            width = 12
        else:
            width = max(14, min(40, len(column.header) + 6))
        sheet.column_dimensions[letter].width = width

    money_format = '#,##0.00'
    date_format = 'DD-MMM-YYYY'

    for row in report.rows():
        cells = []
        for column in report.columns:
            cell = WriteOnlyCell(sheet, value=_xlsx_value(row[column.key]))
            if column.kind == "money" and cell.value is not None:
                cell.number_format = money_format
            elif column.kind == "date" and cell.value is not None:
                cell.number_format = date_format
            cells.append(cell)
        sheet.append(cells)

    # Totals go on their own sheet so they never get mistaken for a data row
    # by whatever opens the file next.
    totals = report.totals()
    if totals:
        summary = workbook.create_sheet(title="Summary")
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 18
        label_font = Font(bold=True)

        for key, value in totals.items():
            label = WriteOnlyCell(summary, value=key.replace("_", " ").title())
            label.font = label_font

            # Totals arrive as strings so JSON keeps decimal precision. Put
            # real numbers in the sheet, or nobody can sum them in Excel.
            cell = WriteOnlyCell(summary, value=_as_number(value))
            if isinstance(cell.value, float):
                cell.number_format = money_format

            summary.append([label, cell])

    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(handle.name)
    handle.seek(0)

    response = FileResponse(handle, content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = (
        f'attachment; filename="{filename_for(report, "xlsx")}"'
    )
    return response


def export(report, export_format: str):
    """Dispatch to the right writer. Format is validated by the serializer."""
    if export_format == "xlsx":
        return build_xlsx(report)
    return stream_csv(report)
