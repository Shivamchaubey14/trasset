"""
Bulk asset import (FR-10.1).

Three principles shape this:

1. **Validation is not duplicated.** Rows go through ``AssetWriteSerializer``,
   the same serializer the API uses, so import rules and API rules cannot drift
   apart. Only the lookup-by-name step is import-specific.
2. **Names, not ids.** A spreadsheet says "Laptops", not "3". Categories,
   locations, departments and vendors are resolved case-insensitively by name,
   and an unknown name is a row error naming what was not found.
3. **Nothing lands half-imported.** Every row is validated before anything is
   written. By default a single bad row aborts the whole file; pass
   ``partial=True`` to import the good rows and report the rest.
"""
import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.masters.models import Category, Department, Location, Vendor

#: Refuse anything larger — a bigger load belongs in a background job, not a
#: request that will time out behind Nginx.
MAX_ROWS = 5000


@dataclass
class ColumnSpec:
    """One importable column."""

    header: str
    field: str
    required: bool = False
    lookup: str = ""      # masters model to resolve by name
    kind: str = "text"    # text | date | decimal | integer
    help_text: str = ""
    example: str = ""


COLUMNS = [
    ColumnSpec("Name", "name", required=True, help_text="What the asset is.",
               example="Dell Latitude 5440"),
    ColumnSpec("Category", "category_id", required=True, lookup="category",
               help_text="Must match an existing category name.",
               example="Laptops"),
    ColumnSpec("Asset Tag", "asset_tag",
               help_text="Leave blank and Trasset generates one.", example=""),
    ColumnSpec("Serial Number", "serial_number", example="SN-DL5440-0091"),
    ColumnSpec("Manufacturer", "manufacturer", example="Dell"),
    ColumnSpec("Model Number", "model_number", example="Latitude 5440"),
    ColumnSpec("Location", "location_id", lookup="location",
               help_text="Existing location name.", example="Head Office — Mumbai"),
    ColumnSpec("Department", "department_id", lookup="department",
               help_text="Existing department name.", example="Information Technology"),
    ColumnSpec("Vendor", "vendor_id", lookup="vendor",
               help_text="Existing vendor name.", example="Dell Technologies India"),
    ColumnSpec("Purchase Date", "purchase_date", kind="date",
               help_text="YYYY-MM-DD or DD/MM/YYYY.", example="2026-01-15"),
    ColumnSpec("Purchase Cost", "purchase_cost", kind="decimal", example="78000.00"),
    ColumnSpec("Salvage Value", "salvage_value", kind="decimal", example="8000.00"),
    ColumnSpec("Useful Life Years", "useful_life_years", kind="integer", example="4"),
    ColumnSpec("Depreciation Method", "depreciation_method",
               help_text="straight_line or declining_balance", example="straight_line"),
    ColumnSpec("Warranty Expiry", "warranty_expiry", kind="date", example="2029-01-15"),
    ColumnSpec("Description", "description", example=""),
    ColumnSpec("Notes", "notes", example=""),
]

HEADER_TO_SPEC = {spec.header.lower(): spec for spec in COLUMNS}

LOOKUP_MODELS = {
    "category": Category,
    "location": Location,
    "department": Department,
    "vendor": Vendor,
}

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%m/%d/%Y")


class ImportError_(Exception):
    """The file itself is unusable — wrong type, empty, or too big."""


@dataclass
class RowResult:
    """What happened to one row."""

    row_number: int
    ok: bool
    errors: dict = field(default_factory=dict)
    asset_tag: str = ""
    name: str = ""
    #: Coerced payload, carried from validation to commit so a real import does
    #: not parse and validate every row twice.
    payload: dict = field(default_factory=dict, repr=False)

    def as_dict(self):
        return {
            "row": self.row_number,
            "ok": self.ok,
            "name": self.name,
            "asset_tag": self.asset_tag,
            "errors": self.errors,
        }


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------
def read_rows(uploaded_file) -> list[dict]:
    """Return a list of ``{header: value}`` dicts from a CSV or XLSX upload."""
    name = (uploaded_file.name or "").lower()

    if name.endswith(".csv"):
        rows = _read_csv(uploaded_file)
    elif name.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(uploaded_file)
    else:
        raise ImportError_(
            "Upload a .csv or .xlsx file. Download the template if you need "
            "the right column headers."
        )

    if not rows:
        raise ImportError_("That file has no data rows.")
    if len(rows) > MAX_ROWS:
        raise ImportError_(
            f"That file has {len(rows)} rows; the limit is {MAX_ROWS} per import. "
            f"Split it into smaller files."
        )
    return rows


def _read_csv(uploaded_file) -> list[dict]:
    raw = uploaded_file.read()
    # utf-8-sig strips the BOM Excel writes, which would otherwise corrupt the
    # first header and make "Name" unmatchable.
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError_("That file's text encoding could not be read. Save it as UTF-8 CSV.")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ImportError_("That file has no header row.")

    return [
        {(key or "").strip(): (value or "").strip()
         for key, value in row.items() if key}
        for row in reader
    ]


def _read_xlsx(uploaded_file) -> list[dict]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a variety of errors
        raise ImportError_("That file could not be opened as a spreadsheet.") from exc

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise ImportError_("That spreadsheet is empty.") from None

    headers = [(str(cell).strip() if cell is not None else "") for cell in header]

    output = []
    for values in rows:
        if all(value is None or str(value).strip() == "" for value in values):
            continue  # blank spacer row
        row = {}
        for index, header_name in enumerate(headers):
            if not header_name:
                continue
            value = values[index] if index < len(values) else None
            row[header_name] = _stringify(value)
        output.append(row)

    workbook.close()
    return output


def _stringify(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


# ---------------------------------------------------------------------------
# Coercion
# ---------------------------------------------------------------------------
def _parse_date(value: str):
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"'{value}' is not a date Trasset recognises (try YYYY-MM-DD).")


def _parse_decimal(value: str):
    # Spreadsheets love thousands separators and currency signs.
    cleaned = value.replace(",", "").replace("₹", "").replace("$", "").strip()
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        raise ValueError(f"'{value}' is not a number.") from None


def _parse_integer(value: str):
    try:
        return int(Decimal(value.replace(",", "").strip()))
    except (InvalidOperation, ValueError):
        raise ValueError(f"'{value}' is not a whole number.") from None


class LookupCache:
    """
    Resolve master names to ids, case-insensitively, one query per model.

    A 500-row import referencing five categories should not issue 500 lookups.
    """

    def __init__(self):
        self._maps = {}

    def _map_for(self, lookup: str) -> dict:
        if lookup not in self._maps:
            model = LOOKUP_MODELS[lookup]
            self._maps[lookup] = {
                row["name"].strip().lower(): row["id"]
                for row in model.objects.values("id", "name")
            }
        return self._maps[lookup]

    def resolve(self, lookup: str, value: str):
        mapping = self._map_for(lookup)
        found = mapping.get(value.strip().lower())
        if found is None:
            raise ValueError(
                f"No {lookup} called '{value}'. Create it under Master Data first, "
                f"or correct the spelling."
            )
        return found

    def known(self, lookup: str) -> list[str]:
        return sorted(LOOKUP_MODELS[lookup].objects.values_list("name", flat=True))


def coerce_row(row: dict, lookups: LookupCache) -> tuple[dict, dict]:
    """
    Turn one spreadsheet row into serializer input.

    :returns: ``(payload, errors)`` — errors keyed by column header.
    """
    payload = {}
    errors = {}

    for header, raw in row.items():
        spec = HEADER_TO_SPEC.get(header.strip().lower())
        if spec is None:
            continue  # unknown column, ignored rather than fatal

        value = (raw or "").strip()
        if not value:
            continue

        try:
            if spec.lookup:
                payload[spec.field] = lookups.resolve(spec.lookup, value)
            elif spec.kind == "date":
                payload[spec.field] = _parse_date(value)
            elif spec.kind == "decimal":
                payload[spec.field] = _parse_decimal(value)
            elif spec.kind == "integer":
                payload[spec.field] = _parse_integer(value)
            else:
                payload[spec.field] = value
        except ValueError as exc:
            errors.setdefault(spec.header, []).append(str(exc))

    for spec in COLUMNS:
        if spec.required and spec.field not in payload and spec.header not in errors:
            errors.setdefault(spec.header, []).append("This column is required.")

    return payload, errors


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
def validate_rows(rows, request=None) -> list[RowResult]:
    """
    Validate every row without writing anything.

    Uses ``AssetWriteSerializer``, so import validation is the API's validation.
    """
    from apps.assets.serializers import AssetWriteSerializer

    lookups = LookupCache()
    results = []
    seen_tags = set()
    seen_serials = set()

    for index, row in enumerate(rows, start=2):  # row 1 is the header
        payload, errors = coerce_row(row, lookups)

        # Duplicates within the file itself — the serializer only checks the DB.
        tag = (payload.get("asset_tag") or "").strip().lower()
        if tag:
            if tag in seen_tags:
                errors.setdefault("Asset Tag", []).append(
                    "This tag appears more than once in the file."
                )
            seen_tags.add(tag)

        serial = (payload.get("serial_number") or "").strip().lower()
        if serial:
            if serial in seen_serials:
                errors.setdefault("Serial Number", []).append(
                    "This serial number appears more than once in the file."
                )
            seen_serials.add(serial)

        if not errors:
            serializer = AssetWriteSerializer(
                data=payload, context={"request": request} if request else {}
            )
            if not serializer.is_valid():
                for field_name, messages in serializer.errors.items():
                    header = _header_for_field(field_name)
                    errors.setdefault(header, []).extend(
                        [str(message) for message in messages]
                    )

        results.append(RowResult(
            row_number=index,
            ok=not errors,
            errors=errors,
            name=payload.get("name", "") or row.get("Name", ""),
            payload=payload,
        ))

    return results


def _header_for_field(field_name: str) -> str:
    for spec in COLUMNS:
        if spec.field == field_name:
            return spec.header
    return field_name.replace("_", " ").title()


@transaction.atomic
def commit_rows(results, request=None) -> list[RowResult]:
    """
    Write the valid rows.

    Reuses the payload cached during validation rather than parsing and
    coercing every row a second time. Runs inside a transaction, so an
    unexpected failure rolls the whole import back rather than leaving half a
    file imported.

    Note on cost: each row still costs roughly four queries, one per foreign
    key the serializer resolves. That is the price of validating imports with
    the same serializer the API uses, which is a trade worth making — and
    ``MAX_ROWS`` bounds the total. If imports get much larger they belong in a
    background job.
    """
    from apps.assets.serializers import AssetWriteSerializer

    for result in results:
        if not result.ok:
            continue

        serializer = AssetWriteSerializer(
            data=result.payload, context={"request": request} if request else {}
        )
        serializer.is_valid(raise_exception=True)
        asset = serializer.save()
        result.asset_tag = asset.asset_tag

    return results


def summarise(results, committed: bool, partial: bool) -> dict:
    valid = [row for row in results if row.ok]
    invalid = [row for row in results if not row.ok]

    return {
        "total_rows": len(results),
        "valid_rows": len(valid),
        "invalid_rows": len(invalid),
        "created": len([row for row in valid if row.asset_tag]) if committed else 0,
        "committed": committed,
        "partial": partial,
        "rows": [row.as_dict() for row in results],
    }


def build_template_csv() -> str:
    """
    A CSV template with headers and one worked example.

    The example uses **this installation's** master data where it exists, so the
    row someone downloads actually imports. A template whose example row is
    rejected teaches the wrong thing on first contact.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # One query per master model, not one per column.
    available = {
        lookup: list(model.objects.values_list("name", flat=True).order_by("name"))
        for lookup, model in LOOKUP_MODELS.items()
    }

    def pick(spec):
        """
        Prefer a master whose name matches the column's own example, so the
        template reads sensibly — a laptop filed under Laptops, not under
        whichever category happens to sort first.
        """
        names = available.get(spec.lookup) or []
        if not names:
            return spec.example

        wanted = spec.example.strip().lower()
        for name in names:
            if name.strip().lower() == wanted:
                return name
        return names[0]

    example = [pick(spec) if spec.lookup else spec.example for spec in COLUMNS]

    writer.writerow([spec.header for spec in COLUMNS])
    writer.writerow(example)

    return output.getvalue()
