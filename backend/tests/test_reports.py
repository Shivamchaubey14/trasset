"""Reports and exports — FR-11.3, FR-11.4, FR-10.2."""
import csv
import io
from datetime import date, timedelta
from decimal import Decimal

from apps.assets.constants import AssetStatus
from apps.assets.models import Asset, AssetAssignment
from apps.assets.services import assignment as assignment_service
from apps.audit.services import suspend
from apps.maintenance.constants import MaintenanceStatus, MaintenanceType
from apps.maintenance.models import MaintenanceRecord
from apps.masters.models import Category, Department, Location, Vendor

from .base import TrassetAPITestCase

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ReportTestCase(TrassetAPITestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.laptops = Category.objects.create(name="Laptops")
        cls.vehicles = Category.objects.create(name="Vehicles")
        cls.office = Location.objects.create(name="Head Office")
        cls.depot = Location.objects.create(name="Depot")
        cls.it = Department.objects.create(name="IT")
        cls.ops = Department.objects.create(name="Operations")
        cls.vendor = Vendor.objects.create(name="Dell India")

        with suspend():
            cls.laptop = Asset.objects.create(
                name="Dell Latitude", category=cls.laptops, location=cls.office,
                department=cls.it, vendor=cls.vendor,
                purchase_date=date.today() - timedelta(days=400),
                purchase_cost=Decimal("80000.00"), salvage_value=Decimal("8000.00"),
                useful_life_years=4,
            )
            cls.van = Asset.objects.create(
                name="Delivery Van", category=cls.vehicles, location=cls.depot,
                department=cls.ops, vendor=cls.vendor,
                purchase_date=date.today() - timedelta(days=900),
                purchase_cost=Decimal("600000.00"), salvage_value=Decimal("60000.00"),
                useful_life_years=8,
            )
            # An asset with no cost — excluded from the depreciation report.
            cls.chair = Asset.objects.create(
                name="Office Chair", category=cls.laptops, location=cls.office,
                purchase_cost=Decimal("0.00"), useful_life_years=5,
            )

    def url(self, key):
        return f"/api/v1/reports/{key}/"

    def run_report(self, key, user=None, **query):
        self.login(user or self.manager)
        response = self.client.get(self.url(key), query)
        self.assertEqual(response.status_code, 200, response.content[:400])
        return response.json()["data"]


class ReportIndexTests(ReportTestCase):
    def test_index_lists_every_report(self):
        self.login(self.manager)
        response = self.client.get("/api/v1/reports/")
        self.assertEqual(response.status_code, 200)

        keys = {row["key"] for row in response.json()["data"]}
        self.assertEqual(
            keys,
            {"asset-register", "depreciation", "maintenance-cost", "assignment"},
        )

    def test_index_describes_the_columns(self):
        self.login(self.manager)
        rows = self.client.get("/api/v1/reports/").json()["data"]
        register = next(row for row in rows if row["key"] == "asset-register")

        self.assertTrue(register["columns"])
        self.assertIn("Asset Tag", [c["header"] for c in register["columns"]])

    def test_an_unknown_report_is_404(self):
        self.login(self.manager)
        response = self.client.get(self.url("does-not-exist"))
        self.assertEqual(response.status_code, 404)
        self.assertIn("is not a report", response.json()["message"])


class AssetRegisterReportTests(ReportTestCase):
    def test_lists_every_asset(self):
        data = self.run_report("asset-register")
        self.assertEqual(data["count"], 3)
        self.assertEqual(data["title"], "Asset Register")

    def test_totals_reconcile_with_the_rows(self):
        """SRS §11.4 — report totals must agree with the register."""
        data = self.run_report("asset-register")
        self.assertEqual(data["totals"]["assets"], 3)
        self.assertEqual(Decimal(data["totals"]["purchase_cost"]),
                         Decimal("680000.00"))

    def test_soft_deleted_assets_are_excluded(self):
        self.chair.delete()
        data = self.run_report("asset-register")
        self.assertEqual(data["count"], 2)

    def test_filter_by_category(self):
        data = self.run_report("asset-register", category=self.vehicles.id)
        self.assertEqual(data["count"], 1)
        self.assertEqual(data["results"][0]["name"], "Delivery Van")

    def test_filter_by_department(self):
        data = self.run_report("asset-register", department=self.it.id)
        self.assertEqual(data["count"], 1)

    def test_filter_by_location(self):
        data = self.run_report("asset-register", location=self.depot.id)
        self.assertEqual(data["count"], 1)

    def test_filter_by_date_range(self):
        data = self.run_report(
            "asset-register",
            date_from=(date.today() - timedelta(days=500)).isoformat(),
        )
        self.assertEqual(data["count"], 1)   # laptop only; van is older, chair has no date

    def test_an_inverted_date_range_is_rejected(self):
        self.login(self.manager)
        response = self.client.get(self.url("asset-register"), {
            "date_from": date.today().isoformat(),
            "date_to": (date.today() - timedelta(days=10)).isoformat(),
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn("date_to", response.json()["errors"])

    def test_pagination(self):
        data = self.run_report("asset-register", page_size=2)
        self.assertEqual(len(data["results"]), 2)
        self.assertEqual(data["total_pages"], 2)

    def test_page_size_is_capped(self):
        self.login(self.manager)
        response = self.client.get(self.url("asset-register"), {"page_size": 99999})
        self.assertEqual(response.status_code, 400)


class DepreciationReportTests(ReportTestCase):
    def test_excludes_assets_with_no_cost(self):
        """A zero-cost asset has nothing to depreciate and is only noise."""
        data = self.run_report("depreciation")
        self.assertEqual(data["count"], 2)
        names = {row["name"] for row in data["results"]}
        self.assertNotIn("Office Chair", names)

    def test_accumulated_plus_book_value_equals_cost(self):
        data = self.run_report("depreciation")
        for row in data["results"]:
            with self.subTest(asset=row["asset_tag"]):
                self.assertEqual(
                    Decimal(row["accumulated_depreciation"]) + Decimal(row["current_value"]),
                    Decimal(row["purchase_cost"]),
                )

    def test_annual_charge_is_reported(self):
        data = self.run_report("depreciation", category=self.laptops.id)
        row = data["results"][0]
        # (80000 - 8000) / 4 = 18000
        self.assertEqual(Decimal(row["annual_depreciation"]), Decimal("18000.00"))

    def test_totals_include_accumulated_depreciation(self):
        data = self.run_report("depreciation")
        totals = data["totals"]
        self.assertEqual(
            Decimal(totals["purchase_cost"]) - Decimal(totals["current_value"]),
            Decimal(totals["accumulated_depreciation"]),
        )


class MaintenanceCostReportTests(ReportTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        with suspend():
            MaintenanceRecord.objects.create(
                asset=cls.laptop, type=MaintenanceType.REPAIR,
                status=MaintenanceStatus.COMPLETED,
                scheduled_date=date.today() - timedelta(days=10),
                completed_date=date.today() - timedelta(days=8),
                cost_estimate=Decimal("2000.00"), actual_cost=Decimal("3200.00"),
                technician="Farhan Q.",
            )
            MaintenanceRecord.objects.create(
                asset=cls.van, type=MaintenanceType.PREVENTIVE,
                status=MaintenanceStatus.SCHEDULED,
                scheduled_date=date.today() + timedelta(days=5),
                cost_estimate=Decimal("15000.00"),
            )

    def test_lists_every_job(self):
        data = self.run_report("maintenance-cost")
        self.assertEqual(data["count"], 2)

    def test_variance_is_calculated(self):
        data = self.run_report("maintenance-cost", category=self.laptops.id)
        row = data["results"][0]
        self.assertEqual(Decimal(row["variance"]), Decimal("1200.00"))

    def test_totals_cover_estimate_and_actual(self):
        totals = self.run_report("maintenance-cost")["totals"]
        self.assertEqual(totals["jobs"], 2)
        self.assertEqual(totals["completed"], 1)
        self.assertEqual(Decimal(totals["estimated_cost"]), Decimal("17000.00"))
        self.assertEqual(Decimal(totals["actual_cost"]), Decimal("3200.00"))

    def test_filter_by_department_reaches_through_the_asset(self):
        data = self.run_report("maintenance-cost", department=self.ops.id)
        self.assertEqual(data["count"], 1)
        self.assertEqual(data["results"][0]["asset_name"], "Delivery Van")

    def test_an_unstarted_job_shows_a_blank_actual(self):
        data = self.run_report("maintenance-cost", category=self.vehicles.id)
        self.assertEqual(data["results"][0]["actual_cost"], "")


class AssignmentReportTests(ReportTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.employee.department = cls.it
        cls.employee.save(update_fields=["department"])

        with suspend():
            assignment_service.assign(cls.laptop, user=cls.employee, actor=cls.manager,
                                      notes="Issued for onboarding")
            assignment_service.checkin(cls.laptop, actor=cls.manager)

    def test_lists_every_movement(self):
        data = self.run_report("assignment")
        self.assertEqual(data["count"], 2)

    def test_newest_first(self):
        data = self.run_report("assignment")
        self.assertEqual(data["results"][0]["action"], "Check-in")
        self.assertEqual(data["results"][1]["action"], "Check-out")

    def test_totals_split_checkouts_and_checkins(self):
        totals = self.run_report("assignment")["totals"]
        self.assertEqual(totals["movements"], 2)
        self.assertEqual(totals["checkouts"], 1)
        self.assertEqual(totals["checkins"], 1)

    def test_department_comes_from_the_person(self):
        data = self.run_report("assignment", department=self.it.id)
        self.assertEqual(data["count"], 2)
        self.assertEqual(data["results"][0]["department"], "IT")

    def test_notes_are_carried_through(self):
        data = self.run_report("assignment")
        checkout = data["results"][1]
        self.assertEqual(checkout["notes"], "Issued for onboarding")


class ExportTests(ReportTestCase):
    """FR-10.2 — CSV and XLSX. PDF is deferred to v1.1."""

    def download(self, key, export_format, user=None, **query):
        self.login(user or self.manager)
        response = self.client.get(self.url(key),
                                   {**query, "export": export_format})
        self.assertEqual(response.status_code, 200)
        return response

    def csv_rows(self, response):
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        return list(csv.reader(io.StringIO(content)))

    def test_csv_download_headers(self):
        response = self.download("asset-register", "csv")
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn("trasset-asset-register-", response["Content-Disposition"])

    def test_csv_contains_a_header_row_and_every_asset(self):
        rows = self.csv_rows(self.download("asset-register", "csv"))
        self.assertEqual(rows[0][0], "Asset Tag")
        self.assertEqual(len(rows), 4)   # header + 3 assets

    def test_csv_starts_with_a_bom_so_excel_reads_utf8(self):
        response = self.download("asset-register", "csv")
        content = b"".join(response.streaming_content)
        self.assertTrue(content.startswith(b"\xef\xbb\xbf"))

    def test_csv_respects_filters(self):
        rows = self.csv_rows(
            self.download("asset-register", "csv", category=self.vehicles.id)
        )
        self.assertEqual(len(rows), 2)   # header + van

    def test_csv_formats_money_and_dates_plainly(self):
        rows = self.csv_rows(self.download("asset-register", "csv"))
        headers = rows[0]
        row = rows[1]

        cost = row[headers.index("Purchase Cost")]
        self.assertRegex(cost, r"^\d+\.\d{2}$")

        purchased = row[headers.index("Purchase Date")]
        if purchased:
            self.assertRegex(purchased, r"^\d{4}-\d{2}-\d{2}$")

    def test_xlsx_download_headers(self):
        response = self.download("asset-register", "xlsx")
        self.assertEqual(response["Content-Type"], XLSX_TYPE)
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_xlsx_is_a_real_workbook(self):
        from openpyxl import load_workbook

        response = self.download("depreciation", "xlsx")
        content = b"".join(response.streaming_content)

        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook[workbook.sheetnames[0]]

        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        self.assertIn("Asset Tag", header)
        self.assertEqual(sheet.max_row, 3)   # header + 2 costed assets

    def test_xlsx_has_a_separate_summary_sheet(self):
        """Totals must not sit where they could be mistaken for data."""
        from openpyxl import load_workbook

        response = self.download("asset-register", "xlsx")
        workbook = load_workbook(io.BytesIO(b"".join(response.streaming_content)))
        self.assertIn("Summary", workbook.sheetnames)

    def test_summary_totals_are_numbers_not_text(self):
        """Otherwise nobody can sum them in Excel."""
        from openpyxl import load_workbook

        response = self.download("asset-register", "xlsx")
        workbook = load_workbook(io.BytesIO(b"".join(response.streaming_content)))
        summary = workbook["Summary"]

        values = {row[0].value: row[1].value for row in summary.iter_rows()}

        # A whole amount round-trips as int rather than float, which is fine —
        # what matters is that it is numeric and not a string.
        self.assertIsInstance(values["Purchase Cost"], (int, float))
        self.assertNotIsInstance(values["Purchase Cost"], str)
        self.assertIsInstance(values["Assets"], int)

    def test_every_report_exports_in_both_formats(self):
        for key in ("asset-register", "depreciation", "maintenance-cost", "assignment"):
            for export_format in ("csv", "xlsx"):
                with self.subTest(report=key, format=export_format):
                    response = self.download(key, export_format)
                    self.assertEqual(response.status_code, 200)

    def test_pdf_is_not_offered(self):
        """Deferred to v1.1 — it must fail clearly, not silently return CSV."""
        self.login(self.manager)
        response = self.client.get(self.url("asset-register"), {"export": "pdf"})
        self.assertEqual(response.status_code, 400)
        self.assertIn("export", response.json()["errors"])

    def test_an_empty_report_still_produces_a_valid_file(self):
        rows = self.csv_rows(
            self.download("asset-register", "csv",
                          date_from=(date.today() + timedelta(days=30)).isoformat())
        )
        self.assertEqual(len(rows), 1)   # header only


class ReportPermissionTests(ReportTestCase):
    def test_every_role_can_run_reports(self):
        """Auditors especially — read-only compliance is the point of the role."""
        for user in self.users.values():
            with self.subTest(role=user.role_name):
                self.login(user)
                response = self.client.get(self.url("asset-register"))
                self.assertEqual(response.status_code, 200)

    def test_auditors_can_export(self):
        self.login(self.auditor)
        response = self.client.get(self.url("asset-register"), {"export": "csv"})
        self.assertEqual(response.status_code, 200)

    def test_requires_authentication(self):
        self.assertEqual(self.client.get(self.url("asset-register")).status_code, 401)
        self.assertEqual(self.client.get("/api/v1/reports/").status_code, 401)


class ReportQueryTests(ReportTestCase):
    def test_report_query_count_does_not_grow_with_rows(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        self.login(self.manager)

        with CaptureQueriesContext(connection) as first:
            self.client.get(self.url("asset-register"), {"page_size": 1})

        with suspend():
            for index in range(20):
                Asset.objects.create(
                    name=f"Extra {index}", category=self.laptops,
                    location=self.office, department=self.it, vendor=self.vendor,
                    purchase_cost=Decimal("1000.00"), useful_life_years=4,
                    purchase_date=date.today(),
                )

        with CaptureQueriesContext(connection) as second:
            self.client.get(self.url("asset-register"), {"page_size": 100})

        self.assertLessEqual(len(second.captured_queries),
                             len(first.captured_queries))
